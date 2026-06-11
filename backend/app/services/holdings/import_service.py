from __future__ import annotations

import csv
import io
import json
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from uuid import uuid4

from app.schemas.financial import HOLDING_ASSET_CLASSES, Holding
from app.services.llm.llm_client import LLMUnavailable
from app.services.llm.ollama_client import OllamaClient
from app.services.llm.schemas import LLMRequest

logger = logging.getLogger(__name__)

# Patterns ranked by specificity. When scanning a header row, the column with
# the HIGHEST-priority match per field wins (e.g. "CURRENT MARKET VALUE" beats
# the more generic "VALUE AT COST"). Higher priority number = more specific.
HEADER_PATTERNS: dict[str, list[tuple[int, re.Pattern]]] = {
    "name": [
        (10, re.compile(r"\b(scheme[\s_]?name|fund[\s_]?name|stock[\s_]?name|instrument[\s_]?name|asset[\s_]?name|holding[\s_]?name|security[\s_/]?scheme)\b", re.I)),
        (5, re.compile(r"\b(security|investment|name|description|particulars|scrip)\b", re.I)),
    ],
    "symbol": [
        (10, re.compile(r"\b(isin|tradingsymbol|exchange[\s_]?symbol)\b", re.I)),
        (5, re.compile(r"\b(symbol|ticker|nse|bse|scrip[\s_]?code)\b", re.I)),
    ],
    "schemeCode": [
        (10, re.compile(r"\b(amfi[\s_]?code|scheme[\s_]?code)\b", re.I)),
        (5, re.compile(r"\b(folio[\s_]?code)\b", re.I)),
    ],
    "units": [
        (10, re.compile(r"\b(quantity[\s_/]?units|holding[\s_]?units|qty)\b", re.I)),
        (5, re.compile(r"\b(units|quantity|shares|grams|gms?|nominal|balance)\b", re.I)),
    ],
    "currentValue": [
        # Strong: explicitly says "current" or "market value" or "present value"
        (20, re.compile(r"\b(current[\s_]?market[\s_]?value|current[\s_]?value|present[\s_]?value|market[\s_]?value|mkt[\s_]?val(?:ue)?|market[\s_]?val|position[\s_]?value|holdings?[\s_]?value|valuation)\b", re.I)),
        # Medium: a value column that's clearly NOT cost
        (10, re.compile(r"\b(cur(?:rent)?[\s_]?val|latest[\s_]?value)\b", re.I)),
        # Weak: generic "value" / "amount" — only used if nothing better
        (3, re.compile(r"^(?!.*\bcost\b)(?!.*\binvested\b).*\b(value|amount|holdings?)\b", re.I)),
    ],
    "valueAtCost": [
        (20, re.compile(r"\b(value[\s_]?at[\s_]?cost|cost[\s_]?value|invested[\s_]?(?:value|amount)|purchase[\s_]?value)\b", re.I)),
        (10, re.compile(r"\b(cost[\s_]?basis|buy[\s_]?value|acquisition[\s_]?cost)\b", re.I)),
        (5, re.compile(r"^(?!.*current).*\b(cost|invested)\b", re.I)),
    ],
    "assetType": [
        (10, re.compile(r"\b(asset[\s_]?class|asset[\s_]?type|instrument[\s_]?type|security[\s_]?type)\b", re.I)),
        (5, re.compile(r"\b(segment|type|category)\b", re.I)),
    ],
}

# Free-text heuristic: scheme/fund name keywords → mutualFund vs stock vs etf
MF_KEYWORDS = ("fund", "mutual", "sip", "scheme", "growth", "direct", "regular", "div pld", "idcw")
ETF_KEYWORDS = ("etf", "nippon", "bees", "icicipru", "niftybees", "goldbees")
CRYPTO_KEYWORDS = ("btc", "bitcoin", "eth", "ethereum", "sol", "usdt", "doge", "ada", "xrp", "matic")

# Summary/total rows to drop even when they parse cleanly.
SUMMARY_NAME_RE = re.compile(r"^\s*(grand\s*total|sub\s*total|net\s*total|total|portfolio\s*value|summary)\s*$", re.I)


@dataclass
class ParseResult:
    holdings: list[Holding]
    unmapped_rows: int
    warnings: list[str]


def parse_holdings_file(path: Path, file_type: str) -> ParseResult:
    text = ""
    rows: list[list[str]] = []
    if file_type == "xlsx":
        rows = _read_xlsx_rows(path)
    elif file_type == "csv":
        text = path.read_text(encoding="utf-8", errors="ignore")
        rows = list(csv.reader(io.StringIO(text)))
    else:
        return ParseResult(holdings=[], unmapped_rows=0, warnings=[f"Unsupported file type: {file_type}"])

    return _rows_to_holdings(rows)


def _read_xlsx_rows(path: Path) -> list[list[str]]:
    from openpyxl import load_workbook

    # Some broker-generated XLSX files (e.g. HDFC Securities) don't expose data
    # in read_only mode. Fall back to full load if read_only returns empty.
    def _collect(workbook) -> list[list[str]]:
        sheets = list(workbook.worksheets)
        # If there's an "All" / "Summary" sheet that's the union of the others,
        # use only that to avoid duplicate rows.
        primary = [s for s in sheets if str(s.title).strip().lower() in {"all", "summary", "portfolio", "holdings"}]
        target_sheets = primary if primary else sheets
        out: list[list[str]] = []
        for sheet in target_sheets:
            for raw in sheet.iter_rows(values_only=True):
                out.append(["" if cell is None else str(cell).strip() for cell in raw])
        return out

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        rows = _collect(wb)
        if any(any(c for c in r) for r in rows):
            return rows
    except Exception:  # noqa: BLE001
        pass
    wb = load_workbook(path, data_only=True)
    return _collect(wb)


def _rows_to_holdings(rows: list[list[str]]) -> ParseResult:
    warnings: list[str] = []
    header_idx, column_map = _detect_header(rows)
    holdings: list[Holding] = []
    unmapped = 0
    non_empty_rows = sum(1 for r in rows if _row_has_content(r))
    logger.info(
        "Holdings parse: %d total rows, %d non-empty. Header row=%d columns=%s",
        len(rows), non_empty_rows, header_idx,
        {k: rows[header_idx][v] if header_idx >= 0 and v < len(rows[header_idx]) else "?" for k, v in column_map.items()},
    )

    if header_idx >= 0 and "name" in column_map and "currentValue" in column_map:
        for row in rows[header_idx + 1:]:
            if not _row_has_content(row):
                continue
            name = _cell(row, column_map.get("name"))
            value = _cell_number(row, column_map.get("currentValue"))
            if not name or value <= 0 or SUMMARY_NAME_RE.match(name):
                unmapped += 1
                continue
            units = _cell_number(row, column_map.get("units"))
            value_at_cost = _cell_number(row, column_map.get("valueAtCost"))
            symbol = _cell(row, column_map.get("symbol"))
            scheme_code = _cell(row, column_map.get("schemeCode"))
            asset_type_raw = _cell(row, column_map.get("assetType"))
            asset_class = _infer_asset_class(name, asset_type_raw, symbol, scheme_code, units)
            holdings.append(Holding(
                id=f"upload-{uuid4().hex[:10]}",
                assetClass=asset_class,
                name=name[:160],
                symbol=symbol[:32] if symbol else "",
                schemeCode=scheme_code[:32] if scheme_code else "",
                units=units,
                currentValue=value,
                valueAtCost=value_at_cost,
                hasSip=False,
                sipAmount=0,
                source="upload",
                lastPricedAt="",
            ))

    # Fall back to LLM if the columnar parser found nothing OR very little.
    columnar_recovery = len(holdings) / max(non_empty_rows, 1)
    if not holdings or (non_empty_rows >= 4 and columnar_recovery < 0.3):
        llm_holdings, llm_note = _llm_extract_holdings(rows)
        if llm_holdings:
            holdings = llm_holdings
            unmapped = max(non_empty_rows - len(holdings), 0)
            warnings.append("Used AI to interpret your statement (non-standard format).")
        elif llm_note:
            warnings.append(llm_note)

    if not holdings:
        warnings.append("No holdings rows could be extracted. Try uploading the broker's original XLSX, or add holdings manually.")
    return ParseResult(holdings=holdings, unmapped_rows=unmapped, warnings=warnings)


def _llm_extract_holdings(rows: list[list[str]]) -> tuple[list[Holding], str]:
    """Ask the local LLM to extract holdings from a non-standard sheet.

    Returns (holdings, error_note). On success, error_note is "". On failure,
    holdings is empty and error_note explains why so the user sees something
    actionable.
    """
    sample = _render_rows_for_llm(rows, max_rows=80, max_cells=12, max_cell_chars=60)
    if not sample.strip():
        return [], ""

    allowed_classes = sorted(c for c in HOLDING_ASSET_CLASSES if c not in {"cash", "epfPpf"})
    prompt = (
        "You are extracting investment holdings from a brokerage / portfolio statement that was uploaded as a spreadsheet.\n"
        "The rows below come from the file as-is (CSV-style). Headers may be in any row, missing, or split across rows. "
        "Identify each individual holding (stock, mutual fund, ETF, crypto, bond, NPS, FD, gold, silver, real estate). "
        "Ignore total/summary rows, blank rows, page headers, footnotes, and disclaimers.\n\n"
        f"Allowed assetClass values: {', '.join(allowed_classes)}\n"
        "Guidance:\n"
        "- name: the scheme/stock/instrument name (no AMC suffixes like 'Direct Growth' if redundant; keep the full name otherwise).\n"
        "- units: shares for stocks, units for MFs, coins for crypto, grams for gold/silver. 0 if not present.\n"
        "- currentValue: market value in INR today. Required > 0. Remove ₹, commas, parentheses.\n"
        "- valueAtCost: invested amount / purchase value / cost basis in INR (often labelled 'Value at Cost', 'Invested', 'Cost'). 0 if not present.\n"
        "- symbol: NSE/BSE ticker for stocks/ETFs (e.g. RELIANCE), CoinGecko symbol for crypto (BTC/ETH). Empty otherwise.\n"
        "- schemeCode: 6-digit AMFI scheme code for mutual funds if present. Empty otherwise.\n\n"
        "Return STRICT JSON only, no markdown, no commentary. Schema:\n"
        '{ "holdings": [ { "name": "...", "assetClass": "...", "units": 0, "currentValue": 0, "valueAtCost": 0, "symbol": "", "schemeCode": "" } ] }\n\n'
        "Statement rows:\n"
        f"{sample}\n"
    )

    try:
        client = OllamaClient()
        if not client.is_reachable():
            return [], "AI fallback unavailable — Ollama is not running. Start it or use a standard XLSX format."
        from app.core.config import settings
        model = settings.llm_model_extraction or settings.llm_model_fast or settings.llm_model or "qwen3:8b"
        response = client.generate(
            LLMRequest(
                task="summarize",  # reuse a generic task slot; routing here is cosmetic
                prompt=prompt,
                model=model,
                expect_json=True,
                timeout_seconds=120,
                metadata={"num_ctx": 8192, "num_predict": 2048},
            )
        )
    except LLMUnavailable as exc:
        logger.warning("LLM holdings extraction unavailable: %s", exc)
        return [], "AI fallback could not run. Try a standard XLSX format or add holdings manually."
    except Exception as exc:  # noqa: BLE001 — defensive
        logger.exception("LLM holdings extraction crashed")
        return [], f"AI fallback failed: {exc}"

    if not response.ok or not response.text:
        return [], "AI fallback returned no usable output."

    parsed = _safe_json(response.text)
    if not isinstance(parsed, dict):
        return [], "AI fallback returned malformed output."
    raw_list = parsed.get("holdings")
    if not isinstance(raw_list, list):
        return [], "AI fallback returned no holdings."

    holdings: list[Holding] = []
    rejected = 0
    for item in raw_list:
        if not isinstance(item, dict):
            rejected += 1
            continue
        name = str(item.get("name") or "").strip()
        current_value = _coerce_number(item.get("currentValue"))
        if not name or current_value <= 0:
            rejected += 1
            continue
        asset_class_raw = str(item.get("assetClass") or "other").strip()
        asset_class = asset_class_raw if asset_class_raw in HOLDING_ASSET_CLASSES else "other"
        units = _coerce_number(item.get("units"))
        value_at_cost = _coerce_number(item.get("valueAtCost"))
        symbol = str(item.get("symbol") or "").strip()[:32].upper()
        scheme_code = str(item.get("schemeCode") or "").strip()[:32]
        holdings.append(Holding(
            id=f"upload-{uuid4().hex[:10]}",
            assetClass=asset_class,
            name=name[:160],
            symbol=symbol,
            schemeCode=scheme_code,
            units=units,
            currentValue=current_value,
            valueAtCost=value_at_cost,
            hasSip=False,
            sipAmount=0,
            source="upload",
            lastPricedAt="",
        ))
    if not holdings and rejected:
        logger.warning("LLM returned %d items but none passed validation. Raw: %s", rejected, response.text[:600])
        return [], f"AI fallback parsed {rejected} rows but couldn't extract usable name + value. Try a different format."
    return holdings, ""


def _coerce_number(raw) -> float:
    """Parse a number that may arrive as int/float/str — strips ₹, $, commas, parens."""
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s:
        return 0.0
    cleaned = re.sub(r"[₹$,\s]", "", s)
    cleaned = cleaned.replace("(", "-").replace(")", "")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _render_rows_for_llm(rows: list[list[str]], max_rows: int, max_cells: int, max_cell_chars: int) -> str:
    out_lines: list[str] = []
    for row in rows[:max_rows]:
        trimmed = [str(c or "").strip()[:max_cell_chars] for c in row[:max_cells]]
        if not any(trimmed):
            continue
        # CSV-quote cells that contain commas/quotes
        cells = []
        for cell in trimmed:
            if any(ch in cell for ch in (",", '"', "\n")):
                escaped = cell.replace('"', '""')
                cells.append(f'"{escaped}"')
            else:
                cells.append(cell)
        out_lines.append(",".join(cells))
    return "\n".join(out_lines)


def _safe_json(text: str):
    text = text.strip()
    # Strip common code fences
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```\s*$", "", text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # Try to extract the largest JSON object from the response
        match = re.search(r"\{.*\}", text, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(0))
            except json.JSONDecodeError:
                pass
    return None


def _detect_header(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    best_idx = -1
    best_map: dict[str, int] = {}
    for idx, row in enumerate(rows[:30]):  # only scan first 30 rows for a header
        # Score each column for each field; keep the highest-scoring column per field.
        scores: dict[str, tuple[int, int]] = {}  # field -> (best_score, col_idx)
        for col_idx, cell in enumerate(row):
            cell_str = str(cell or "").strip()
            if not cell_str:
                continue
            for field, patterns in HEADER_PATTERNS.items():
                for priority, pattern in patterns:
                    if pattern.search(cell_str):
                        existing = scores.get(field)
                        if not existing or priority > existing[0]:
                            scores[field] = (priority, col_idx)
                        break  # don't double-score this cell against weaker patterns
        column_map = {field: col_idx for field, (_, col_idx) in scores.items()}
        if "name" in column_map and "currentValue" in column_map and len(column_map) > len(best_map):
            best_idx = idx
            best_map = column_map
    return best_idx, best_map


def _row_has_content(row: Iterable[str]) -> bool:
    return any(str(cell or "").strip() for cell in row)


def _cell(row: list[str], idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    return str(row[idx] or "").strip()


def _cell_number(row: list[str], idx: int | None) -> float:
    raw = _cell(row, idx)
    if not raw:
        return 0.0
    cleaned = re.sub(r"[₹$,\s]", "", raw)
    cleaned = cleaned.replace("(", "-").replace(")", "")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _infer_asset_class(name: str, asset_type_raw: str, symbol: str, scheme_code: str, units: float) -> str:
    lower_name = name.lower()
    lower_type = asset_type_raw.lower() if asset_type_raw else ""

    if any(k in lower_name for k in CRYPTO_KEYWORDS) or "crypto" in lower_type:
        return "crypto"
    if scheme_code or any(k in lower_name for k in MF_KEYWORDS) or "mutual" in lower_type or "mf" in lower_type:
        if any(k in lower_name for k in ETF_KEYWORDS) or "etf" in lower_type:
            return "etf"
        return "mutualFund"
    if any(k in lower_name for k in ETF_KEYWORDS) or "etf" in lower_type:
        return "etf"
    if "bond" in lower_type or "bond" in lower_name or "ncd" in lower_name:
        return "bond"
    if "nps" in lower_type or "nps" in lower_name:
        return "nps"
    if "fd" in lower_type or "fixed deposit" in lower_name or "fixed dep" in lower_type:
        return "fd"
    if "gold" in lower_type or "sgb" in lower_name or "gold" in lower_name:
        return "gold"
    if "silver" in lower_type or "silver" in lower_name:
        return "silver"
    if "real estate" in lower_type or "property" in lower_type:
        return "realEstate"
    # Default: if it has a ticker symbol or whole-number-ish units, treat as a stock
    if symbol or (units and units == int(units)):
        return "stock"
    return "other"
